# This script will load a File from CES, rename and obtain country restrictions.
# a 'Log' worksheet will be appended to the Excel file with these restrictions.
# 
# Last modifications:
# - Openpyxl to append Log to excel (12/08/22)
# - reflected China CICPA changes (24/08/22)
#
#
#


if __name__ == "__main__":           
    import pandas as pd
    import os
    import sys
    old_filename = input("Enter old Filename (eg. 'compliance_detail'): ")
    company_name = input("Enter company name (eg. 'Company Inc.'):")
    target_filename = company_name + "_CES compliance export.xlsx"

    print(os.getcwd())
    print(target_filename)

    if (old_filename != target_filename):
        os.rename(old_filename+".xlsx", target_filename)

    
    df = pd.read_excel(target_filename)

    #df = pd.read_excel("Bain Capital Holdings, LP_CES compliance export.xlsx")

    def results(re,rules):
        if re:
            return rules
        else:
            return "NOT"+" "+rules
            
    AICPA_mask = df['Independence Standard'] == "AICPA" 
    def AICPA_check():
        rules = "Subject to AICPA independence rules."
        res=False
        res |= any(AICPA_mask)
        return results(res, rules)

    GIP_mask = df['Independence Standard'] == "GIP" 
    def GIP_check():
        rules = "Subject to GIP independence rules."
        res=False
        res |= any(GIP_mask)
        return results(res, rules)

    SEC_mask = df['Independence Standard'] == "SEC" 
    def SEC_check():
        rules = "Subject to SEC independence rules."
        res=False
        res |= any(SEC_mask)
        if res:
            return rules + " contact CoE Hub leader"
        else:
            return results(res,rules)

    Belgium_mask = (df['Restricting Entity Country'] == "Belgium") & (df['Restriction Reason'] == "The entity is an audit client")
    def belgium_check():
        rules = "Subject to Belgian independence rules."
        res=False
        res |= any(Belgium_mask)
        return results(res,rules)

    # UPDATE: Old guidance on CICPA has since been deprecated.
    #
    # china_CICPA_mask = df["Independence Standard Reason"].str.contains("Required by Global Independence Policy and subject to China CICPA rules")
    # china_territories = ["China"] #,"Hong Kong","Macao","Macau"]
    # china_audit_mask = (df["Restricting Entity Country"].apply(lambda x: x in china_territories) & (df["Deliverable Type"] == "Audit"))
    # restriction_status_china_mask = df["Restricting Entity Listed status"].apply(lambda x: True if x=="Yes" else False)
    # china_if_mask_1 = (~china_CICPA_mask) & china_audit_mask & restriction_status_china_mask
    # restriction_reason_china_mask = (df["Restriction Reason"] == "The entity is an audit client.") | (df["Restriction Reason"] == "The audit client has direct or indirect control over the entity.")
    # china_if_mask_2 = china_audit_mask & (~china_CICPA_mask) & ~restriction_status_china_mask & restriction_reason_china_mask
    # china_non_audit_mask = (df["Restricting Entity Country"].apply(lambda x: x in china_territories) & ~(df["Deliverable Type"] == "Audit"))
    # china_if_mask_3 = (~china_CICPA_mask) & china_non_audit_mask & (df["Restriction Reason"] == "The entity is a (non-audit) assurance client")
    
    
    # New China guidance
    #
    china_territories = ["China"] #,"Hong Kong","Macao","Macau"]
    china_audit_mask = (df["Restricting Entity Country"].apply(lambda x: x in china_territories) & (df["Deliverable Type"] == "Audit"))
    restriction_status_china_mask = df["Restricting Entity Listed status"].apply(lambda x: True if x=="Yes" else False)

    china_if_mask_1 = china_audit_mask & restriction_status_china_mask
    
    restriction_reason_china_mask = (df["Restriction Reason"] == "The entity is an audit client.") | (df["Restriction Reason"] == "The audit client has direct or indirect control over the entity.")
    china_if_mask_2 = china_audit_mask & ~restriction_status_china_mask & restriction_reason_china_mask
    
    china_non_audit_mask = (df["Restricting Entity Country"].apply(lambda x: x in china_territories) & (df["Deliverable Type"] == "Non-Audit Assurance"))
    china_if_mask_3 =china_non_audit_mask & (df["Restriction Reason"] == "The entity is a (non-audit) assurance client")
        
    
    
    def china_check():
        rules = "Subject to Chinese independence rules."
        res=False
        res |= any(china_if_mask_1) 
        res |= any(china_if_mask_2)
        res |= any(china_if_mask_3)
        if res:
            return rules + " contact CoE Hub leader"
        else:
            return results(res,rules)

    norway_mask = (df['Restricting Entity Country'] == "Norway") & (df["Deliverable Type"] == "Audit")
    def norway_check():
        rules = "Subject to Norwegian independence rules."
        res=False
        res |= any(norway_mask)   
        return results(res,rules)

    sweden_mask = (df['Restricting Entity Country'] == "Sweden") & (df["Deliverable Type"] == "Audit")
    def sweden_check():
        rules = "Subject to Swedish independence rules."
        res=False
        res |= any(sweden_mask)   
        return results(res,rules)

    saudi_audit_mask = (df["Restricting Entity Country"] == "Saudi Arabia") & (df["Deliverable Type"] == "Audit") 
    saudi_if_mask_1 = (df['Restriction Reason'] == "The entity is an audit client.")
    saudi_if_mask_2 = (df['Restriction Reason'] == "The audit client has direct or indirect control over the entity.") & ~(df["Restricting Entity Supplemental Information"].fillna("None").apply(lambda x: x.casefold()).str.contains("sovereign"))
    def saudi_check():
        rules = "Subject to Saudi independence rules."
        res=False
        res |= any(saudi_audit_mask&saudi_if_mask_1)  
        res |= any(saudi_audit_mask&saudi_if_mask_2) 
        if res:
            return rules + " contact CoE Hub leader"
        else:
            return results(res,rules)

    dependencies = ["Isle of Man","Jersey", "Guernsey"]
    uk_dependencies_mask = (df["Restricting Entity Country"].apply(lambda x: x in dependencies))
    uk_mask = (df['Restricting Entity Country'] == "United Kingdom") & (df["Deliverable Type"] == "Audit")
    def uk_check():
        rules = "Subject to UK independence rules."
        res=False
        res |= any(uk_dependencies_mask)  
        res |= any(uk_mask) 
        if any(uk_dependencies_mask):
            return "UK Dependencies rules may apply, contact UK independence office"    
        else:
            return results(res,rules)

    Aus_audit_mask = (df['Restricting Entity Country'] == "Australia") & (df["Deliverable Type"] == "Audit")
    def australia_check(australia = False):
        rules = "Subject to Australian independence rules."
        res=False
        if australia:
            res |= any(Aus_audit_mask)   
        return results(res,rules)

    india_mask = (df['Restricting Entity Country'] == "India")
    def india_check(india = False):
        rules = "Subject to Indian independence rules."
        res=False
        if india:
            res |= any(india_mask)   
        return results(res,rules)

    def all_check(australia = False, india = False):
        country_functions = (AICPA_check(), 
        GIP_check(),
            SEC_check(), 
            australia_check(australia), 
            belgium_check(),
            china_check(),
                india_check(india),
            norway_check(),
            saudi_check(),              
            sweden_check(), 
            uk_check()
            )
        print("\n")     
        for i in country_functions:
            print(i,"\n")          
    all_check()
    
    from openpyxl import load_workbook
    
    wb = load_workbook(target_filename)
    if "log" not in wb.sheetnames:
        wb.create_sheet('log')
    active_sheet = wb["log"]

    country_functions = (AICPA_check(), GIP_check(),
            SEC_check(), 
            australia_check(), 
            belgium_check(),
            china_check(),
            india_check(),
            norway_check(),
            saudi_check(),              
            sweden_check(), 
            uk_check()
            )

    for row in range(1,len(country_functions)+1):
        active_sheet.cell(row=row, column=1, value="{0}".format(country_functions[row-1]))
    
    wb.save(target_filename)
    

    end = input("Run another search? (Y/N)") 
    if (end == "Y") or (end == "y") or (end == "Yes") or (end == "yes") :
        os.startfile(__file__)
        sys.exit()


    
